"""
§25 EXCEL EXPORT — 17 sheet бүхий workbook үүсгэнэ.

Web application-д энэ нь ExcelJS-ээр хийгддэг
(`src/services/export/excel-export.service.ts`). Энэ Python хувилбар нь
ИЖИЛ бүтэц, ИЖИЛ багана, ИЖИЛ тоог гаргах бөгөөд CLI-аар шууд ажиллана.

Форматлалт (§25):
  • форматтай толгой (өнгө, тод, хөлдөөсөн)
  • autofilter
  • freeze panes
  • тоон формат (мянгатын таслал, хувь, бутархай)
  • нөхцөлт форматлалт (эрсдэлийн өнгө §26)
  • нийлбэрийн мөр
  • баганын өргөн автоматаар

⚠️ §28 — эх өгөгдөлд байхгүй утгыг "N/A" эсвэл шалтгаантайгаар харуулна,
   тоо ЗОХИОХГҮЙ.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Sequence

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── §26 ӨНГӨНИЙ СИСТЕМ ────────────────────────────────────────────────
COLORS = {
    "header": "1E293B",       # хар хөх — толгой
    "headerText": "FFFFFF",
    "critical": "FEE2E2",     # улаан — нөөц дуусах
    "lowStock": "FFEDD5",     # улбар шар — нөөц багассан
    "excess": "F3E8FF",       # ягаан — илүүдэл
    "stagnant": "E5E7EB",     # хар саарал — хөдөлгөөнгүй
    "slowMoving": "DBEAFE",   # цэнхэр — удаан эргэлт
    "healthy": "DCFCE7",      # ногоон — зохистой
    "total": "F1F5F9",
}

ABC_XYZ_FILL = {
    "AX": "DCFCE7", "AY": "DCFCE7", "AZ": "FFEDD5",
    "BX": "DCFCE7", "BY": "FEF9C3", "BZ": "FFEDD5",
    "CX": "ECFCCB", "CY": "FFEDD5", "CZ": "FEE2E2",
}

STATUS_FILL = {
    "STOCKOUT_RISK": COLORS["critical"],
    "LOW_STOCK": COLORS["lowStock"],
    "OVERSTOCK": COLORS["excess"],
    "NO_MOVEMENT": COLORS["stagnant"],
    "SLOW_MOVING": COLORS["slowMoving"],
    "OPTIMAL": COLORS["healthy"],
}

STATUS_LABEL = {
    "STOCKOUT_RISK": "Нөөц дуусах эрсдэлтэй",
    "LOW_STOCK": "Нөөц багассан",
    "OVERSTOCK": "Хэт их нөөцтэй",
    "NO_MOVEMENT": "Хөдөлгөөнгүй",
    "SLOW_MOVING": "Удаан эргэлттэй",
    "OPTIMAL": "Зохистой",
}

DECISION_LABEL = {
    "TRANSFER": "Шилжүүлэх",
    "NEW_PURCHASE": "Шинээр худалдан авах",
    "STOP_PURCHASE": "Худалдан авалт зогсоох",
    "MONITOR": "Хяналтад байлгах",
    "PROMOTION": "Борлуулалт идэвхжүүлэх",
}

NA = "N/A"

FMT_INT = "#,##0"
FMT_DEC1 = "#,##0.0"
FMT_DEC2 = "#,##0.00"
FMT_MONEY = "#,##0"
FMT_PCT = "0.0%"

THIN = Side(style="thin", color="CBD5E1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


@dataclass
class Column:
    header: str
    key: str
    width: int = 14
    fmt: str | None = None
    getter: Callable[[Any], Any] | None = None

    def value(self, row: Any) -> Any:
        if self.getter is not None:
            return self.getter(row)
        if isinstance(row, dict):
            return row.get(self.key)
        return getattr(row, self.key, None)


def _write_sheet(
    wb: Workbook,
    title: str,
    columns: Sequence[Column],
    rows: Sequence[Any],
    *,
    subtitle: str | None = None,
    total_keys: Sequence[str] = (),
    fill_by: Callable[[Any], str | None] | None = None,
    empty_note: str = "Өгөгдөл байхгүй",
):
    """Нэг sheet бичих — форматлалт бүрэн."""
    ws = wb.create_sheet(title[:31])
    start = 1

    if subtitle:
        ws.cell(row=1, column=1, value=subtitle).font = Font(italic=True, size=9, color="475569")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(columns)))
        start = 3

    # Толгой
    for i, col in enumerate(columns, start=1):
        cell = ws.cell(row=start, column=i, value=col.header)
        cell.font = Font(bold=True, color=COLORS["headerText"], size=10)
        cell.fill = PatternFill("solid", fgColor=COLORS["header"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = col.width
    ws.row_dimensions[start].height = 30

    # Өгөгдөл
    if not rows:
        ws.cell(row=start + 1, column=1, value=empty_note).font = Font(italic=True, color="94A3B8")
    for r, row in enumerate(rows, start=start + 1):
        fill = fill_by(row) if fill_by else None
        for i, col in enumerate(columns, start=1):
            cell = ws.cell(row=r, column=i, value=col.value(row))
            cell.border = BORDER
            if col.fmt:
                cell.number_format = col.fmt
            if fill:
                cell.fill = PatternFill("solid", fgColor=fill)

    # Нийлбэрийн мөр
    if rows and total_keys:
        tr = start + len(rows) + 1
        ws.cell(row=tr, column=1, value="НИЙТ").font = Font(bold=True)
        for i, col in enumerate(columns, start=1):
            cell = ws.cell(row=tr, column=i)
            cell.fill = PatternFill("solid", fgColor=COLORS["total"])
            cell.font = Font(bold=True)
            cell.border = BORDER
            if col.key in total_keys:
                letter = get_column_letter(i)
                cell.value = f"=SUM({letter}{start + 1}:{letter}{start + len(rows)})"
                cell.number_format = col.fmt or FMT_INT

    # Autofilter + freeze
    if rows:
        last = get_column_letter(len(columns))
        ws.auto_filter.ref = f"A{start}:{last}{start + len(rows)}"
    ws.freeze_panes = ws.cell(row=start + 1, column=1)
    return ws


def _kv_sheet(wb: Workbook, title: str, sections: list[tuple[str, list[tuple[str, Any, str | None]]]]):
    """Түлхүүр-утга хэлбэрийн хураангуй sheet."""
    ws = wb.create_sheet(title[:31])
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 60

    r = 1
    for section_title, items in sections:
        cell = ws.cell(row=r, column=1, value=section_title)
        cell.font = Font(bold=True, color=COLORS["headerText"], size=11)
        cell.fill = PatternFill("solid", fgColor=COLORS["header"])
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        r += 1
        for label, value, note in items:
            ws.cell(row=r, column=1, value=label).font = Font(size=10)
            c = ws.cell(row=r, column=2, value=value)
            c.font = Font(bold=True, size=10)
            if isinstance(value, (int, float)):
                c.number_format = FMT_MONEY
            if note:
                ws.cell(row=r, column=3, value=note).font = Font(size=9, color="64748B")
            r += 1
        r += 1
    ws.freeze_panes = "A2"
    return ws


def build_workbook(data: dict, out_path: Path) -> Path:
    """
    17 sheet бүхий workbook үүсгэнэ.

    `data` нь `python/export/collect.py`-ийн цуглуулсан бүтэц.
    """
    wb = Workbook()
    wb.remove(wb.active)

    meta = data["meta"]
    inv_rows = data["inventory"]
    abc_rows = data["abcXyz"]
    benchmarks = data["benchmarks"]
    recs = data["recommendations"]
    transfers = data["transfers"]
    quality = data["quality"]

    # ── 1. Dashboard Summary ──
    _kv_sheet(wb, "1.Dashboard Summary", [
        ("ТООЦООЛЛЫН ПАРАМЕТР", [
            ("Тооцооны сар", meta["calculationMonth"], "Энэ сар дундажид ОРОХГҮЙ"),
            ("Ашигласан сарууд", ", ".join(meta["periods"]), f"{len(meta['periods'])} бүтэн сар"),
            ("Тайлан үүсгэсэн", datetime.now().strftime("%Y-%m-%d %H:%M"), None),
            ("Эх файл", meta["sourceFile"], None),
        ]),
        ("ХАМРАХ ХҮРЭЭ", [
            ("Нийт SKU", meta["skuCount"], None),
            ("Байршил", meta["locationCount"], "3 агуулах + 7 эмийн сан"),
            ("Байрлал (SKU × байршил)", len(inv_rows), None),
        ]),
        ("БОРЛУУЛАЛТ", [
            ("Нийт борлуулалтын дүн (өртгөөр)", meta["totalSalesValue"],
             "⚠️ Эх өгөгдөлд ОРЛОГО байхгүй — энэ нь COGS"),
            ("Нийт борлуулалтын тоо", meta["totalSalesQty"], None),
        ]),
        ("НӨӨЦ", [
            ("Нийт нөөцийн тоо", meta["totalStockQty"], None),
            ("Нөөцийн өртөг", meta["totalStockValue"], None),
            ("Нийт дутагдал", meta["totalShortage"], None),
            ("Дутагдлын өртөг", meta["totalShortageValue"], None),
            ("Нийт илүүдэл", meta["totalExcess"], None),
            ("Илүүдлийн өртөг", meta["totalExcessValue"], None),
        ]),
        ("АШИГ (§7)", [
            ("Gross Profit", NA, meta["marginReason"]),
            ("Gross Margin %", NA, meta["marginReason"]),
        ]),
        ("ШИЙДВЭР", [
            ("Шилжүүлэх санал", len(transfers), f"{meta['totalTransferQty']:,} ширхэг"),
            ("Шинээр худалдан авах", meta["purchaseSkuCount"],
             f"{meta['totalPurchaseQty']:,} ширхэг"),
            ("Худалдан авалт зогсоох", meta["stopPurchaseCount"], None),
        ]),
        ("ҮНИЙН ХЯНАЛТ", [
            ("Benchmark хийсэн SKU", meta["benchmarkedProducts"], None),
            ("Олон эх сурвалжтай SKU", meta["multiSourceProducts"], None),
            ("Боломжит хэмнэлт", meta["totalPotentialSaving"],
             "Одоогийн худалдан авалтыг хамгийн бага үнээр авсан бол"),
            ("MARGIN RISK SKU", meta["marginRiskProducts"], None),
        ]),
        ("НӨӨЦИЙН ТӨЛӨВ", [
            (STATUS_LABEL.get(k, k), v, None)
            for k, v in sorted(meta["byStatus"].items(), key=lambda kv: -kv[1])
        ]),
    ])

    # ── 2. SKU Analysis ──
    _write_sheet(wb, "2.SKU Analysis", [
        Column("SKU код", "product_code", 12),
        Column("Бүтээгдэхүүн", "product_name", 38),
        Column("ABC", "abc", 6),
        Column("XYZ", "xyz", 6),
        Column("ABCXYZ", "abc_xyz", 9),
        Column("Борлуулалтын дүн", "sales_value", 18, FMT_MONEY),
        Column("Эзлэх %", "sales_share", 10, FMT_PCT),
        Column("Хуримтлагдсан %", "cumulative_share", 14, FMT_PCT),
        Column("Дундаж сарын тоо", "average_monthly_qty", 16, FMT_DEC2),
        Column("StdDev", "std_dev", 12, FMT_DEC2),
        Column("CV", "cv", 10, FMT_DEC2, lambda r: r.cv if r.cv is not None else NA),
        Column("Борлуулалттай сар", "months_with_sales", 16, FMT_INT),
        Column("Төлөв", "inventory_status", 14,
               getter=lambda r: "Хөдөлгөөнгүй" if r.inventory_status == "NO_MOVEMENT" else "Идэвхтэй"),
    ], abc_rows,
        subtitle="ABC = борлуулалтын МӨНГӨН дүнгээр · XYZ = тоо хэмжээний хэлбэлзлээр (STDEV.P)",
        total_keys=("sales_value",),
        fill_by=lambda r: ABC_XYZ_FILL.get(r.abc_xyz))

    # ── 3. ABCXYZ Analysis ──
    matrix = data["matrix"]
    _write_sheet(wb, "3.ABCXYZ Analysis", [
        Column("ABCXYZ", "abcXyz", 10),
        Column("SKU тоо", "skuCount", 10, FMT_INT),
        Column("Борлуулалтын дүн", "salesValue", 18, FMT_MONEY),
        Column("Эзлэх %", "salesShare", 10, FMT_PCT),
        Column("Борлуулалтын тоо", "salesQty", 16, FMT_DEC1),
        Column("Одоогийн нөөц", "currentStock", 15, FMT_DEC1),
        Column("Зохистой нөөц", "recommendedStock", 15, FMT_DEC1),
        Column("Эрсдэлтэй байрлал", "riskCount", 16, FMT_INT),
    ], matrix,
        subtitle="⭐ 9 хосолсон ангилал — дараагийн бүх нөөцийн тооцооллын үндэс",
        total_keys=("skuCount", "salesValue", "salesQty", "currentStock",
                    "recommendedStock", "riskCount"),
        fill_by=lambda r: ABC_XYZ_FILL.get(r["abcXyz"]))

    # ── 4. Inventory Balance ──
    balance_rows = data["balanceSummary"]
    _write_sheet(wb, "4.Inventory Balance", [
        Column("Төлөв", "labelMn", 26),
        Column("Байрлалын тоо", "count", 14, FMT_INT),
        Column("Эзлэх %", "share", 10, FMT_PCT),
        Column("Тоо хэмжээ", "quantity", 16, FMT_DEC1),
        Column("Өртөг", "value", 18, FMT_MONEY),
    ], balance_rows,
        subtitle="НӨӨЦИЙН ТЭНЦВЭР — төлөв тус бүрийн тоо, эзлэх хувь, тоо хэмжээ, өртөг",
        total_keys=("count", "quantity", "value"),
        fill_by=lambda r: STATUS_FILL.get(r["code"]))

    # ── 5. Recommended Stock ──
    inv_cols = [
        Column("SKU код", "product_code", 12, getter=lambda r: r.position.product_code),
        Column("Бүтээгдэхүүн", "product_name", 34, getter=lambda r: r.position.product_name),
        Column("Байршил", "location_code", 11, getter=lambda r: r.position.location_code),
        Column("Төрөл", "location_type", 12,
               getter=lambda r: "Агуулах" if r.position.location_type == "WAREHOUSE" else "Эмийн сан"),
        Column("ABCXYZ", "abc_xyz", 9, getter=lambda r: r.position.abc_xyz),
        Column("Дундаж/сар", "avg", 13, FMT_DEC2,
               lambda r: r.position.average_monthly_sales),
        Column("Зорилтот хоног", "target_days", 14, FMT_INT, lambda r: r.balance.target_days),
        Column("Зорилтот сар", "target_months", 13, FMT_DEC2, lambda r: r.balance.target_months),
        Column("Зохистой нөөц", "recommended", 15, FMT_DEC1,
               lambda r: r.balance.recommended_stock),
        Column("Одоогийн нөөц", "current", 15, FMT_DEC1, lambda r: r.balance.current_stock),
        Column("Нөөцийн хоног", "days", 14, FMT_DEC1, lambda r: r.balance.current_stock_days),
        Column("Дутагдал", "shortage", 13, FMT_DEC1, lambda r: r.balance.shortage),
        Column("Илүүдэл", "excess", 13, FMT_DEC1, lambda r: r.balance.excess),
        Column("Төлөв", "status", 22, getter=lambda r: STATUS_LABEL.get(r.stock_status, r.stock_status)),
        Column("Шийдвэр", "decision", 22, getter=lambda r: DECISION_LABEL.get(r.decision, r.decision)),
    ]
    _write_sheet(wb, "5.Recommended Stock", inv_cols, inv_rows,
                 subtitle="Зохистой нөөц = дундаж сарын борлуулалт × (зорилтот хоног ÷ 30)",
                 total_keys=("recommended", "current", "shortage", "excess"),
                 fill_by=lambda r: STATUS_FILL.get(r.stock_status))

    # ── 6. Risk SKU ──
    risk_rows = data["riskRows"]
    _write_sheet(wb, "6.Risk SKU", [
        Column("Priority", "priority", 10),
        Column("SKU код", "productCode", 12),
        Column("Бүтээгдэхүүн", "productName", 34),
        Column("ABCXYZ", "abcXyz", 9),
        Column("Байршил", "locationCode", 11),
        Column("Одоогийн нөөц", "currentStock", 15, FMT_DEC1),
        Column("Нөөцийн хоног", "stockDays", 14, FMT_DEC1),
        Column("Зорилтот хоног", "targetDays", 14, FMT_INT),
        Column("Дутагдал", "shortage", 13, FMT_DEC1),
        Column("Дутагдлын өртөг", "shortageValue", 16, FMT_MONEY),
        Column("Эрсдэл", "risk", 22),
        Column("AI зөвлөмж", "action", 60),
    ], risk_rows,
        subtitle="Эрсдэлийн зэрэг ба нөлөөллийн дүнгээр эрэмбэлэгдсэн",
        total_keys=("shortage", "shortageValue"),
        fill_by=lambda r: COLORS["critical"] if r["priority"] == "CRITICAL"
        else COLORS["lowStock"] if r["priority"] == "HIGH" else None)

    # ── 7. Excess Inventory ──
    excess_rows = [r for r in inv_rows if r.balance.excess > 0]
    _write_sheet(wb, "7.Excess Inventory", [
        Column("SKU код", "c", 12, getter=lambda r: r.position.product_code),
        Column("Бүтээгдэхүүн", "n", 34, getter=lambda r: r.position.product_name),
        Column("ABCXYZ", "a", 9, getter=lambda r: r.position.abc_xyz),
        Column("Байршил", "l", 11, getter=lambda r: r.position.location_code),
        Column("Одоогийн нөөц", "current", 15, FMT_DEC1, lambda r: r.balance.current_stock),
        Column("Зохистой нөөц", "recommended", 15, FMT_DEC1,
               lambda r: r.balance.recommended_stock),
        Column("Илүүдэл", "excess", 13, FMT_DEC1, lambda r: r.balance.excess),
        Column("Илүүдлийн хоног", "excessDays", 15, FMT_DEC1,
               lambda r: max(0.0, r.balance.current_stock_days - r.balance.target_days)),
        Column("Илүүдлийн өртөг", "excessValue", 16, FMT_MONEY,
               lambda r: r.excess_value if r.excess_value is not None else NA),
        Column("Санал", "d", 24, getter=lambda r: DECISION_LABEL.get(r.decision, r.decision)),
    ], sorted(excess_rows, key=lambda r: -(r.excess_value or 0)),
        subtitle="Илүүдлийн хоног = одоогийн нөөцийн хоног − зорилтот хоног",
        total_keys=("excess", "excessValue"),
        fill_by=lambda r: COLORS["excess"])

    # ── 8. Stagnant Inventory ──
    stagnant = data["stagnantRows"]
    _write_sheet(wb, "8.Stagnant Inventory", [
        Column("SKU код", "productCode", 12),
        Column("Бүтээгдэхүүн", "productName", 34),
        Column("Байршил", "locationCode", 11),
        Column("Одоогийн нөөц", "currentStock", 15, FMT_DEC1),
        Column("Нөөцийн өртөг", "stockValue", 16, FMT_MONEY),
        Column("Сүүлийн борлуулалт", "lastSalesPeriod", 18,
               getter=lambda r: r["lastSalesPeriod"] or NA),
        Column("Хэдэн сар зарагдаагүй", "monthsSinceLastSale", 20, FMT_INT,
               lambda r: r["monthsSinceLastSale"] if r["monthsSinceLastSale"] is not None else NA),
        Column("Сүүлийн худалдан авалт", "lastPurchasePeriod", 20,
               getter=lambda r: r["lastPurchasePeriod"] or NA),
        Column("Санал", "recommendation", 26),
    ], stagnant,
        subtitle="⚠️ Эх өгөгдөлд ӨДРИЙН огноо байхгүй — огнооны оронд САР (YYYY-MM). "
                 "Хөдөлгөөнгүй = lookback хугацаанд дундаж борлуулалт 0.",
        total_keys=("currentStock", "stockValue"),
        fill_by=lambda r: COLORS["stagnant"])

    # ── 9. Slow Moving ──
    slow = [r for r in inv_rows if r.position.xyz == "Z"]
    _write_sheet(wb, "9.Slow Moving", [
        Column("SKU код", "c", 12, getter=lambda r: r.position.product_code),
        Column("Бүтээгдэхүүн", "n", 34, getter=lambda r: r.position.product_name),
        Column("ABCXYZ", "a", 9, getter=lambda r: r.position.abc_xyz),
        Column("Байршил", "l", 11, getter=lambda r: r.position.location_code),
        Column("Дундаж/сар", "avg", 13, FMT_DEC2, lambda r: r.position.average_monthly_sales),
        Column("Одоогийн нөөц", "current", 15, FMT_DEC1, lambda r: r.balance.current_stock),
        Column("Нөөцийн хоног", "days", 14, FMT_DEC1, lambda r: r.balance.current_stock_days),
        Column("Зорилтот хоног", "target", 14, FMT_INT, lambda r: r.balance.target_days),
        Column("Илүүдэл", "excess", 13, FMT_DEC1, lambda r: r.balance.excess),
        Column("Төлөв", "s", 22, getter=lambda r: STATUS_LABEL.get(r.stock_status, r.stock_status)),
    ], sorted(slow, key=lambda r: -r.balance.excess),
        subtitle="XYZ = Z (эрэлт нь маш хэлбэлзэлтэй) бүх байрлал",
        total_keys=("current", "excess"),
        fill_by=lambda r: COLORS["slowMoving"])

    # ── 10. Transfer Recommendation ──
    _write_sheet(wb, "10.Transfer Recommendation", [
        Column("Эрэмбэ", "priority_rank", 8, FMT_INT),
        Column("SKU код", "product_code", 12),
        Column("Бүтээгдэхүүн", "product_name", 34,
               getter=lambda t: data["nameByCode"].get(t.product_code)),
        Column("Хаанаас", "from_location_code", 11),
        Column("Хаашаа", "to_location_code", 11),
        Column("Эх үүсвэрийн илүүдэл", "sourceSurplus", 18, FMT_DEC1,
               lambda t: data["excessByKey"].get((t.product_code, t.from_location_code), 0)),
        Column("Хүлээн авагчийн дутагдал", "destShortage", 20, FMT_DEC1,
               lambda t: data["shortageByKey"].get((t.product_code, t.to_location_code), 0)),
        Column("Шилжүүлэх тоо", "quantity", 14, FMT_INT),
        Column("Үлдэх дутагдал", "remaining", 15, FMT_DEC1,
               lambda t: max(0.0, data["shortageByKey"].get(
                   (t.product_code, t.to_location_code), 0)
                   - data["transferInByKey"].get((t.product_code, t.to_location_code), 0))),
        Column("Дүн", "estimated_value", 15, FMT_MONEY,
               lambda t: t.estimated_value if t.estimated_value is not None else NA),
        Column("Шалтгаан", "reason_mn", 52),
    ], sorted(transfers, key=lambda t: -t.quantity),
        subtitle="БАЙРШИЛ ХООРОНД ШИЛЖҮҮЛЭХ САНАЛ — тоо нь БҮХЭЛ, эх үүсвэрийн илүүдлээс хэтрэхгүй",
        total_keys=("quantity",))

    # ── 11. Purchase Recommendation ──
    purchase_rows = [r for r in inv_rows if r.new_purchase_qty > 0 or r.transfer_in_qty > 0
                     or r.decision == "STOP_PURCHASE"]
    _write_sheet(wb, "11.Purchase Recommendation", [
        Column("SKU код", "c", 12, getter=lambda r: r.position.product_code),
        Column("Бүтээгдэхүүн", "n", 34, getter=lambda r: r.position.product_name),
        Column("ABCXYZ", "a", 9, getter=lambda r: r.position.abc_xyz),
        Column("Байршил", "l", 11, getter=lambda r: r.position.location_code),
        Column("Одоогийн нөөц", "current", 15, FMT_DEC1, lambda r: r.balance.current_stock),
        Column("Зохистой нөөц", "target", 15, FMT_DEC1, lambda r: r.balance.recommended_stock),
        Column("Дутагдал", "shortage", 13, FMT_DEC1, lambda r: r.balance.shortage),
        Column("Шилжүүлэг боломжтой", "avail", 18,
               getter=lambda r: "Тийм" if r.transfer_in_qty > 0 else "Үгүй"),
        Column("Шилжүүлэх тоо", "transfer", 14, FMT_INT, lambda r: r.transfer_in_qty),
        Column("Худалдан авах тоо", "purchase", 16, FMT_INT, lambda r: r.new_purchase_qty),
        Column("Шийдвэр", "d", 24, getter=lambda r: DECISION_LABEL.get(r.decision, r.decision)),
    ], sorted(purchase_rows, key=lambda r: -(r.shortage_value or 0)),
        subtitle="ТАТАН АВАЛТЫН ШИЙДВЭР — эхлээд шилжүүлэг, дараа нь шинэ худалдан авалт",
        total_keys=("shortage", "transfer", "purchase"),
        fill_by=lambda r: STATUS_FILL.get(r.stock_status))

    # ── 12. Purchase Price Control ──
    bench_usable = [b for b in benchmarks if b.min_unit_price is not None]
    _write_sheet(wb, "12.Purchase Price Control", [
        Column("SKU код", "product_code", 12),
        Column("Бүтээгдэхүүн", "product_name", 34),
        Column("Эх сурвалж", "source_count", 11, FMT_INT),
        Column("Хамгийн бага үнэ", "min_unit_price", 16, FMT_DEC2),
        Column("Хамгийн бага эх с.", "min_source_key", 16),
        Column("Хамгийн өндөр үнэ", "max_unit_price", 16, FMT_DEC2),
        Column("Хамгийн өндөр эх с.", "max_source_key", 17),
        Column("Үнийн зөрүү", "price_gap", 14, FMT_DEC2),
        Column("Зөрүү %", "price_gap_pct", 11, FMT_DEC1),
        Column("Зэрэглэл", "gap_severity", 12, getter=lambda b: b.gap_severity or "—"),
        Column("Одоогийн тоо", "current_quantity", 14, FMT_DEC1),
        Column("Одоогийн өртөг", "current_cost", 16, FMT_MONEY),
        Column("Боломжит хэмнэлт", "potential_saving", 17, FMT_MONEY),
        Column("Өртгийн өөрчлөлт %", "price_change_pct", 18, FMT_DEC1,
               lambda b: b.price_change_pct if b.price_change_pct is not None else NA),
        Column("MARGIN RISK", "risk", 13,
               getter=lambda b: "ТИЙМ" if data["marginRiskCodes"] and
               b.product_code in data["marginRiskCodes"] else "—"),
    ], sorted(bench_usable, key=lambda b: -(b.potential_saving or 0)),
        subtitle="⚠️ Эх сурвалж = НИЙЛҮҮЛЭГЧ (эх өгөгдөлд суваг байхгүй). "
                 "Хэмнэлт = одоогийн худалдан авалтыг хамгийн бага үнээр авсан бол.",
        total_keys=("current_cost", "potential_saving"),
        fill_by=lambda b: COLORS["critical"] if b.gap_severity == "CRITICAL"
        else COLORS["lowStock"] if b.gap_severity == "HIGH" else None)

    # ── 13 / 14. Lowest & Highest price TOP 3 ──
    def price_points(kind: str) -> list[dict]:
        out = []
        for b in bench_usable:
            pts = b.lowest_top if kind == "low" else b.highest_top
            for p in pts:
                out.append({
                    "productCode": b.product_code,
                    "productName": b.product_name,
                    "rank": p.lowest_rank if kind == "low" else p.highest_rank,
                    "source": p.dimension_key,
                    "period": p.last_purchase_period,
                    "qty": p.quantity,
                    "amount": p.amount,
                    "unitPrice": p.unit_price,
                })
        return sorted(out, key=lambda x: (x["productCode"], x["rank"]))

    price_cols = [
        Column("SKU код", "productCode", 12),
        Column("Бүтээгдэхүүн", "productName", 34),
        Column("Эрэмбэ", "rank", 8, FMT_INT),
        Column("Эх сурвалж (нийлүүлэгч)", "source", 20),
        Column("Сүүлийн худалдан авалт", "period", 20),
        Column("Тоо", "qty", 12, FMT_DEC1),
        Column("Дүн", "amount", 16, FMT_MONEY),
        Column("Нэгж үнэ", "unitPrice", 14, FMT_DEC2),
    ]
    _write_sheet(wb, "13.Lowest Price TOP 3", price_cols, price_points("low"),
                 subtitle="§3 — SKU бүрийн хамгийн БАГА нэгж үнэтэй TOP 3 эх сурвалж",
                 fill_by=lambda r: COLORS["healthy"] if r["rank"] == 1 else None)
    _write_sheet(wb, "14.Highest Price TOP 3", price_cols, price_points("high"),
                 subtitle="§4 — SKU бүрийн хамгийн ӨНДӨР нэгж үнэтэй TOP 3 эх сурвалж",
                 fill_by=lambda r: COLORS["critical"] if r["rank"] == 1 else None)

    # ── 15. Gross Margin Risk ──
    _write_sheet(wb, "15.Gross Margin Risk", [
        Column("SKU код", "product_code", 12),
        Column("Бүтээгдэхүүн", "product_name", 34),
        Column("Зарах үнэ", "salesPrice", 14, getter=lambda b: NA),
        Column("Худалдан авалтын өртөг", "cost", 20, FMT_DEC2,
               lambda b: b.weighted_avg_unit_price),
        Column("Gross Profit", "gp", 14, getter=lambda b: NA),
        Column("Gross Margin %", "gm", 14, getter=lambda b: NA),
        Column("Хамгийн бага үнэ", "min_unit_price", 16, FMT_DEC2),
        Column("Одоогийн үнэ", "cur", 14, FMT_DEC2,
               lambda b: b.points[0].unit_price if b.points else None),
        Column("Үнийн зөрүү", "price_gap", 14, FMT_DEC2),
        Column("Зөрүү %", "price_gap_pct", 11, FMT_DEC1),
        Column("Боломжит алдагдал", "potential_saving", 18, FMT_MONEY),
        Column("Эрсдэлийн шалтгаан", "reasons", 60,
               getter=lambda b: " · ".join(data["marginRiskReasons"].get(b.product_code, []))),
    ], sorted([b for b in bench_usable if b.product_code in data["marginRiskCodes"]],
              key=lambda b: -(b.potential_saving or 0)),
        subtitle="⚠️ Зарах үнэ / Gross Profit / Gross Margin — ЭХ ӨГӨГДӨЛД БАЙХГҮЙ (N/A). "
                 "Эрсдэлийг үнийн зөрүү ба өртгийн өсөлтөөр илрүүлсэн.",
        total_keys=("potential_saving",),
        fill_by=lambda b: COLORS["critical"])

    # ── 16. AI Recommendation ──
    _write_sheet(wb, "16.AI Recommendation", [
        Column("Priority", "priority", 10),
        Column("Эрсдэл", "risk", 22),
        Column("SKU код", "product_code", 12),
        Column("Бүтээгдэхүүн", "product_name", 30),
        Column("Байршил", "location_code", 11),
        Column("ABCXYZ", "abc_xyz", 9),
        Column("WHY", "reason", 70),
        Column("IMPACT", "impact", 70),
        Column("ACTION", "recommended_action", 70),
        Column("Шилжүүлэх боломж", "transfer_possible", 16,
               getter=lambda r: "Тийм" if r["transfer_possible"] else "Үгүй"),
        Column("Худалдан авалт", "purchase_required", 15,
               getter=lambda r: "Тийм" if r["purchase_required"] else "Үгүй"),
        Column("Зогсоох", "stop_purchase", 10,
               getter=lambda r: "Тийм" if r["stop_purchase"] else "Үгүй"),
        Column("Санал болгох тоо", "recommended_quantity", 16, FMT_INT),
    ], recs,
        subtitle="⚠️ AI нь тооцооллын үр дүнг ӨӨРЧЛӨӨГҮЙ — 'Санал болгох тоо' нь "
                 "engine-ийн бодсон шилжүүлэг/худалдан авалтын тоо.",
        total_keys=("recommended_quantity",),
        fill_by=lambda r: COLORS["critical"] if r["priority"] == "CRITICAL"
        else COLORS["lowStock"] if r["priority"] == "HIGH" else None)

    # ── 17. Data Quality ──
    _write_sheet(wb, "17.Data Quality", [
        Column("Sheet", "sheet", 14),
        Column("Дүрмийн код", "code", 28),
        Column("Хүнд байдал", "severity", 13),
        Column("Тоо", "count", 10, FMT_INT),
        Column("Тайлбар", "message", 60),
    ], quality["issues"],
        subtitle=f"Нийт {quality['total']:,} мөр · VALID {quality['valid']:,} · "
                 f"WARNING {quality['warning']:,} · ERROR {quality['error']:,}",
        total_keys=("count",),
        fill_by=lambda r: COLORS["critical"] if r["severity"] == "ERROR"
        else COLORS["lowStock"])

    # Нөхцөлт форматлалт — дутагдалтай мөрүүд
    ws = wb["5.Recommended Stock"]
    if inv_rows:
        rng = f"L4:L{3 + len(inv_rows)}"
        ws.conditional_formatting.add(rng, CellIsRule(
            operator="greaterThan", formula=["0"],
            fill=PatternFill("solid", bgColor=COLORS["critical"])))

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
