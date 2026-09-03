"""
Excel structure inspector — давтан ажиллуулж болох хэлбэрээр.

docs/01-excel-inspection.md-д байгаа бүх тоо ЭНЭ script-ээс гарсан.
Шинэ файл ирэх бүрд дахин ажиллуулж, баримтыг шинэчилнэ.

Ажиллуулах:
    python python/inspect_excel.py "C:\\path\\to\\Data AI.xlsx"

Windows консол дээр Cyrillic хэвлэхэд:
    set PYTHONIOENCODING=utf-8
"""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
import pandas as pd

# Product code-ыг ХЭЗЭЭ Ч int болгож болохгүй (тэргүүлэх 0 алдагдана)
STRING_COLUMNS = {"Дотоод код"}


def cell_storage_types(path: Path) -> None:
    """Excel дотор утга нь текстээр хадгалагдсан эсэхийг шалгана."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    print("\n=== CELL STORAGE TYPES (мөр 2) ===")
    for ws in wb.worksheets:
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        row2 = next(ws.iter_rows(min_row=2, max_row=2))
        print(f"\n[{ws.title}]")
        for name, cell in zip(header, row2):
            kind = "text" if cell.data_type == "s" else "number" if cell.data_type == "n" else cell.data_type
            print(f"  {name!r:<28} storage={kind}")
    wb.close()


def profile(path: Path) -> None:
    xl = pd.ExcelFile(path)
    print(f"SHEETS: {xl.sheet_names}")

    for sheet in xl.sheet_names:
        df = xl.parse(sheet, dtype={c: str for c in STRING_COLUMNS})
        print("\n" + "#" * 78)
        print(f"SHEET {sheet}: rows={len(df)} cols={len(df.columns)}")
        for col in df.columns:
            s = df[col]
            print(f"\n  col={col!r}")
            print(
                f"     dtype={s.dtype} non_null={s.notna().sum()}/{len(s)} "
                f"unique={s.nunique(dropna=True)}"
            )
            if pd.api.types.is_numeric_dtype(s):
                print(
                    f"     min={s.min()} max={s.max()} "
                    f"neg={int((s < 0).sum())} zero={int((s == 0).sum())}"
                )
            if s.nunique(dropna=True) <= 20:
                print(f"     values={dict(s.value_counts(dropna=False))}")
            else:
                print(f"     sample={list(s.dropna().unique()[:4])}")


def dimension_report(path: Path) -> None:
    """ХХК → Суваг → Төрөл шатлал ба хугацааны хамрах хүрээ."""
    print("\n=== DIMENSION HIERARCHY ===")
    for sheet in ("Purchase", "Sales", "Stock"):
        df = pd.read_excel(path, sheet, dtype={"Дотоод код": str})
        print(f"\n[{sheet}]")
        print(df.groupby(["ХХК", "Төрөл", "Суваг"]).size().to_string())
        ym = df["Он"].astype(str) + "-" + df["Сар"].astype(str).str.zfill(2)
        print(f"  periods: {sorted(ym.unique())}")


def price_semantics(path: Path) -> None:
    """`Өртөг` нь орлого мөн үү, өртөг мөн үү — нэгж үнээр шалгана."""
    print("\n=== PRICE SEMANTICS CHECK ===")
    dt = {"Дотоод код": str}
    pur = pd.read_excel(path, "Purchase", dtype=dt)
    sal = pd.read_excel(path, "Sales", dtype=dt)

    def unit(df, qty, amt, mask=None):
        d = df if mask is None else df[mask]
        return d.groupby("Дотоод код").apply(
            lambda g: g[amt].sum() / g[qty].sum() if g[qty].sum() else float("nan"),
            include_groups=False,
        )

    pu = unit(pur, "Тоо", "ТА НӨАТгүй дүн").rename("purchase_unit")
    su = unit(sal, "Тоо", "Өртөг", sal["Төрөл"] == "ЭХНТ").rename("sales_unit")
    cmp = pd.concat([pu, su], axis=1).dropna()
    ratio = (cmp["sales_unit"] / cmp["purchase_unit"]).median()
    print(f"  median(sales_unit / purchase_unit) = {ratio:.3f}")
    print(
        "  → ~1.00 гэдэг нь Sales.'Өртөг' нь COGS (өртөг), борлуулалтын орлого БИШ гэсэн үг."
    )


def main() -> int:
    if len(sys.argv) < 2:
        print(__doc__)
        return 2
    path = Path(sys.argv[1])
    if not path.exists():
        print(f"Файл олдсонгүй: {path}")
        return 1

    profile(path)
    cell_storage_types(path)
    dimension_report(path)
    price_semantics(path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
